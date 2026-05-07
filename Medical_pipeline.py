"""
================================================================
MEDICAL CLASSIFICATION PIPELINE — FINAL VERSION
================================================================
"""
import pandas as pd, numpy as np, re, unicodedata, warnings, json
from pathlib import Path

warnings.filterwarnings('ignore')
from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split, StratifiedKFold, cross_val_score
from sklearn.metrics import f1_score
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import LabelEncoder

# ====================== CẤU HÌNH ======================
OUT_DIR = Path(r"D:\PipelinesCategory\Pipeline\Category_pipelines\Medical")
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ─────────────────────────────────────────────────────────────
# 0. LOAD DATA
# ─────────────────────────────────────────────────────────────
print("="*70)
print("📂 LOAD DATA - MEDICAL")
print("="*70)

wb = load_workbook('Data_2025_2026_Combined_Cleaned.xlsx', read_only=True)
ws = wb['Medical']
rows = list(ws.iter_rows(values_only=True))
data = [r[:4] for r in rows[1:] if r and any(v is not None for v in r[:4])]

df = pd.DataFrame(data, columns=['category','user_id','product_group','item_name'])
df['item_name'] = df['item_name'].fillna('').astype(str).str.strip()
df['product_group'] = df['product_group'].fillna('').astype(str).str.strip()

print(f"✅ Loaded {len(df):,} rows")

# Label Ground
GROUP_NORM = {
    'Thuốc': 'Thuốc', 'Thuốc ': 'Thuốc',
    'Dụng cụ Y tế': 'Dụng cụ Y tế', 'Dụng cụ y tế': 'Dụng cụ Y tế',
    'Dịch vụ Y tế': 'Dịch vụ Y tế', 'Dịch vụ y tế': 'Dịch vụ Y tế',
    'Thực phẩm chức năng': 'Thực phẩm chức năng',
    'Dung dịch vệ sinh/ Y tế': 'Dung dịch vệ sinh',
    'Dung dịch vệ sinh': 'Dung dịch vệ sinh',
    'Bỉm -Tã': 'Dụng cụ Y tế',
}
df['label_ground'] = df['product_group'].map(GROUP_NORM).fillna('Khác')

# Preprocessing
def preprocess(text):
    if not text or not text.strip(): return ''
    t = unicodedata.normalize('NFC', text).lower()
    t = re.sub(r'\b\d+[\.,]?\d*\s*(?:ml|l|g|gram|kg|miếng|hộp|chai|gói|viên|ống)\b', ' ', t)
    t = re.sub(r'\b\d{8,}\b', ' ', t)
    t = re.sub(r'\b[a-z]{1,3}\d{3,}\b', ' ', t)
    t = re.sub(r'[^\w\s\u00C0-\u024F\u1E00-\u1EFF\-&+/\']', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t

df['clean'] = df['item_name'].apply(preprocess)

# ─────────────────────────────────────────────────────────────
# RULE-BASED
# ─────────────────────────────────────────────────────────────
RULES = [
    ('Thuốc', 1, [r'\b(thuốc|thuoc|siro|hapacol|medrol|prednison|prednisolone|methylprednisolon|augmentin|efferalgan|alpha choay|acemuc)\b']),
    ('Dụng cụ Y tế', 1, [r'\b(khẩu trang|gạc|pmc|urgo|sterimar|dr\.? ?papie|salonpas|que thử|kim tiêm)\b']),
    ('Dịch vụ Y tế', 2, [r'\b(khám|siêu âm|xét nghiệm|chụp xquang|công khám)\b']),
    ('Thực phẩm chức năng', 2, [r'\b(tp bvsk|tpbvsk|enterogermina|livespo|men vi sinh)\b']),
    ('Dung dịch vệ sinh', 3, [r'\b(lactacyd|femfresh|dung dịch vệ sinh)\b']),
]

def rule_classify(text):
    if not text.strip(): return 'Khác', 0.0, ''
    tl = text.lower()
    for label, priority, patterns in sorted(RULES, key=lambda x: x[1]):
        for pat in patterns:
            if re.search(pat, tl, re.IGNORECASE):
                conf = 1.0 if priority == 1 else (0.85 if priority == 2 else 0.75)
                return label, conf, pat
    return 'Khác', 0.0, ''

res = df['clean'].apply(rule_classify)
df['rule_label'] = [r[0] for r in res]
df['rule_conf']  = [r[1] for r in res]
df['rule_match'] = [r[2] for r in res]

df['noise_flag'] = df['item_name'].str.contains('trà sữa', case=False, na=False)
df_noise = df[df['noise_flag']]

# ─────────────────────────────────────────────────────────────
# ML TRAINING (char + word)
# ─────────────────────────────────────────────────────────────
print("\n🤖 Training & Cross-validation...")

df_tr = df[df['label_ground'] != 'Khác'].copy()
le = LabelEncoder()
y = le.fit_transform(df_tr['label_ground'])
X = df_tr['clean'].fillna('')

# Model A - char TF-IDF
pipe_a = Pipeline([
    ('tfidf', TfidfVectorizer(ngram_range=(2,4), analyzer='char_wb', max_features=10000, sublinear_tf=True)),
    ('clf', LogisticRegression(max_iter=500, C=3.0, class_weight='balanced'))
])
cv_a = cross_val_score(pipe_a, X, y, cv=StratifiedKFold(5, shuffle=True, random_state=42), scoring='f1_macro')

# Model B - word TF-IDF
pipe_b = Pipeline([
    ('tfidf', TfidfVectorizer(ngram_range=(1,2), max_features=10000, sublinear_tf=True)),
    ('clf', LogisticRegression(max_iter=500, C=3.0, class_weight='balanced'))
])
cv_b = cross_val_score(pipe_b, X, y, cv=StratifiedKFold(5, shuffle=True, random_state=42), scoring='f1_macro')

# Chọn model tốt hơn
best_pipe = pipe_a if cv_a.mean() >= cv_b.mean() else pipe_b
best_name = 'char TF-IDF + LR' if cv_a.mean() >= cv_b.mean() else 'word TF-IDF + LR'
best_cv = max(cv_a.mean(), cv_b.mean())

best_pipe.fit(X, y)

# Predict toàn bộ dữ liệu
prob = best_pipe.predict_proba(df['clean'].fillna(''))
pred = np.argmax(prob, axis=1)
conf = np.max(prob, axis=1)

df['ml_label'] = le.inverse_transform(pred)
df['ml_conf']  = np.round(conf, 3)

# Final decision
def decide(row):
    if row['label_ground'] != 'Khác':
        return row['label_ground'], 'ground-truth'
    if row['noise_flag']:
        return 'Loại bỏ', 'noise'
    if row['ml_conf'] >= 0.85:
        return row['ml_label'], 'ml-high'
    if row['ml_conf'] >= 0.60:
        return row['ml_label'], 'ml-medium'
    return 'Cần rà soát', 'uncertain'

dec = df.apply(decide, axis=1)
df['pred_label'] = [r[0] for r in dec]
df['pred_source'] = [r[1] for r in dec]

df['flag_review'] = (df['pred_source'] == 'uncertain').astype(int)

# ─────────────────────────────────────────────────────────────
# EXPORT
# ─────────────────────────────────────────────────────────────
cols = ['category','user_id','product_group','item_name','clean',
        'label_ground','rule_label','rule_conf','ml_label','ml_conf',
        'pred_label','pred_source','flag_review','noise_flag']

df[cols].to_csv(OUT_DIR / 'medical_full_labeled.csv', index=False, encoding='utf-8-sig')

df_rft = df[df['pred_source'].isin(['ground-truth', 'ml-high']) & (df['pred_label'] != 'Loại bỏ')]
df_rft[['item_name','clean','pred_label']].to_csv(OUT_DIR / 'medical_ready_for_train.csv', index=False, encoding='utf-8-sig')

df_rev = df[df['pred_source'].isin(['uncertain','ml-medium'])][cols]
df_rev.to_csv(OUT_DIR / 'medical_review_queue.csv', index=False, encoding='utf-8-sig')

# Rules
rules_rows = []
for label, prio, pats in RULES:
    for p in pats:
        rules_rows.append({'label': label, 'priority': prio, 'pattern': p})

pd.DataFrame(rules_rows).to_csv(OUT_DIR / 'medical_keyword_rules_v1.csv', index=False, encoding='utf-8-sig')

# ====================== SUMMARY BOX (GIỐNG COSMETICS) ======================
print(f"""
╔══════════════════════════════════════════════════════╗
║        MEDICAL PIPELINE — SUMMARY                    ║
╠══════════════════════════════════════════════════════╣
║  Tổng items            : {len(df):>6}                    ║
║  CV F1 char TF-IDF+LR  : {cv_a.mean()*100:>6.2f}%                  ║
║  CV F1 word TF-IDF+LR  : {cv_b.mean()*100:>6.2f}%                  ║
║  Best model            : {best_name:<26}  ║
║  Best CV F1            : {best_cv*100:>6.2f}%                  ║
║  Ready-for-train       : {len(df_rft):>6}                    ║
║  Review queue          : {len(df_rev):>6}                    ║
║  Noise removed         : {len(df_noise):>6}                    ║
║  Keyword rules         : {len(rules_rows):>6}                    ║
╚══════════════════════════════════════════════════════╝
""")

print(f"\n🎉 HOÀN THÀNH! Tất cả file đã được lưu vào thư mục:\n{OUT_DIR}")