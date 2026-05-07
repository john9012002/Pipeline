"""
================================================================
COSMETICS CLASSIFICATION PIPELINE — PHA 3
4 sub-groups chính:
  1. Chăm sóc da (Skincare)
  2. Trang điểm (Makeup)
  3. Chăm sóc tóc & thân (Haircare & Body)
  4. Vệ sinh cá nhân (Personal Care)
================================================================
"""
import pandas as pd, numpy as np, re, unicodedata, warnings, json
warnings.filterwarnings('ignore')
from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split, StratifiedKFold, cross_val_score
from sklearn.metrics import f1_score, classification_report
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import LabelEncoder

OUT = 'D:\\EDA\\Category_pipelines\\cosmetics'

# ─────────────────────────────────────────────────────────────
# 0. LOAD DATA
# ─────────────────────────────────────────────────────────────
print("="*60); print("📂 0. LOAD DATA"); print("="*60)
wb = load_workbook('Data_2025_2026_Combined_Cleaned.xlsx', read_only=True)
ws = wb['Cosmetics']
rows = list(ws.iter_rows(values_only=True))
data = [r[:4] for r in rows[1:] if r and any(v is not None for v in r[:4])]
df = pd.DataFrame(data, columns=['category','user_id','product_group','item_name'])
df['item_name']     = df['item_name'].fillna('').astype(str).str.strip()
df['product_group'] = df['product_group'].fillna('').astype(str).str.strip()
print(f"✅ {len(df):,} rows | {df['product_group'].nunique()} product groups")

# ─────────────────────────────────────────────────────────────
# 0b. CHUẨN HÓA NHÃN GỐC → 4 sub-groups
# ─────────────────────────────────────────────────────────────
GROUP_NORM = {
    # Skincare
    'Kem bôi da -Trị Mụn':                   'Skincare',
    'Mặt nạ':                                 'Skincare',
    'Serum':                                  'Skincare',
    'Toner-Nước Hoa Hồng':                    'Skincare',
    'Kem chống nắng - Sữa chống nắng':        'Skincare',
    'Tẩy da chết - BHA':                      'Skincare',
    'Kem dưỡng ẩm - Sữa dưỡng - Sáp':        'Skincare',
    'Sữa rửa mặt':                            'Skincare',
    'Sữa rửa mặt-Gel rửa mặt-Kem':           'Skincare',
    'Tẩy trang':                              'Skincare',
    # Makeup
    'Son môi - Son tint':                     'Makeup',
    'Son':                                    'Makeup',
    'Phấn má - Phấn phủ':                     'Makeup',
    'Che khuyết điểm (concealer)':            'Makeup',
    'Xịt khoá nền':                           'Makeup',
    'Kem lót - Kem nền':                      'Makeup',
    'Giấy thấm dầu':                          'Makeup',
    'Bông mút trang điểm-Dụng cụ làm đẹp':   'Makeup',
    # Haircare & Body
    'Dầu gội':                                'Haircare & Body',
    'Dầu xả':                                 'Haircare & Body',
    'Sữa tắm':                                'Haircare & Body',
    'Sữa Tắm':                                'Haircare & Body',
    'Xà bông':                                'Haircare & Body',
    'Xà bông ':                               'Haircare & Body',
    'Lăn nách':                               'Haircare & Body',
    'Nước hoa':                               'Haircare & Body',
    # Personal Care
    'Kem đánh răng':                          'Personal Care',
    'Bàn chải đánh răng':                     'Personal Care',
    'Nước súc miệng':                         'Personal Care',
    'Băng vệ sinh':                           'Personal Care',
    'Khăn ướt':                               'Personal Care',
    'Dung dịch vệ sinh phụ nữ':              'Personal Care',
}
df['label_ground'] = df['product_group'].map(GROUP_NORM).fillna('Khác')
print("\nNhãn sau chuẩn hóa:")
print(df['label_ground'].value_counts().to_string())

# ─────────────────────────────────────────────────────────────
# 1. PREPROCESSING
# ─────────────────────────────────────────────────────────────
print("\n"+"="*60); print("⚙️  1. PREPROCESSING"); print("="*60)

TYPO = {
    r'\bkdr\b':           'kem đánh răng',
    r'\bbvs\b':           'băng vệ sinh',
    r'\bst\b(?=\s)':      'sữa tắm',
    r'\bdg\b(?=\s)':      'dầu gội',
    r'\bdx\b(?=\s)':      'dầu xả',
    r'\bsrm\b':           'sữa rửa mặt',
    r'\bxbc\b':           'xà bông cục',
    r'\bxnm\b':           'xịt nước mùi',
    r'\blnm\b':           'lăn nách mùi',
    r'\bntt\b':           'nước tẩy trang',
    r'\bspf\b':           'chống nắng spf',
    r'\bkem\s*u\s*toc\b': 'kem ủ tóc',
    r'\bcap\s*am\b':      'cấp ẩm',
    r'\bduong\s*chat\b':  'dưỡng chất',
}
UNIT_RE = re.compile(
    r'\b\d+[\.,]?\d*\s*(?:ml|l|g|gram|kg|tờ|miếng|cái|hộp|tuýp|chai|gói|set|pack|viên)\b',
    re.IGNORECASE)

def preprocess(text):
    if not text or not text.strip(): return ''
    t = unicodedata.normalize('NFC', text).lower()
    for p, r in TYPO.items():
        t = re.sub(p, r, t, flags=re.IGNORECASE)
    t = UNIT_RE.sub(' ', t)
    # Loại mã SKU, barcode, mã màu
    t = re.sub(r'\b\d{8,}\b', ' ', t)
    t = re.sub(r'\b[a-z]{1,3}\d{3,}\b', ' ', t)
    t = re.sub(r'[^\w\s\u00C0-\u024F\u1E00-\u1EFF\-&+/\']', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t

df['clean'] = df['item_name'].apply(preprocess)
print(f"✅ Preprocessing xong {len(df):,} items")
print("\nVí dụ:")
for _, r in df[['item_name','clean']].sample(6, random_state=1).iterrows():
    print(f"  RAW : {r['item_name'][:60]}")
    print(f"  CLEAN: {r['clean'][:60]}\n")

# ─────────────────────────────────────────────────────────────
# BƯỚC 1: RULE-BASED LABELING
# ─────────────────────────────────────────────────────────────
print("="*60); print("📏 BƯỚC 1: RULE-BASED LABELING"); print("="*60)

RULES = [
    # ── PRIORITY 1: Rất rõ ràng ───────────────────────────
    ('Personal Care', 1, [
        r'\bkem\s*đánh\s*răng\b',r'\bkem\s*danh\s*rang\b',r'\bkdr\b',
        r'\bcolgate\b',r'\bsensodyne\b',r'\bsensdyne\b',r'\bpepsodent\b',
        r'\bp/?s\b(?=\s*(kdr|kem|răng))',r'\boral[\s\-]?b\b',
        r'\bbàn\s*chải\s*(đánh\s*răng|răng)?\b',r'\btoothbrush\b',r'\btoothpaste\b',
        r'\bnước\s*súc\s*miệng\b',r'\blisterine\b',r'\bmouthwash\b',
        r'\bbăng\s*vệ\s*sinh\b',r'\bbvs\b',
        r'\bkotex\b',r'\blaurier\b',r'\bdiana\b(?=\s*(bvs|băng))',
        r'\btampon\b',r'\btã\b',r'\bbỉm\b',r'\bmama\b(?=\s*\d)',
        r'\bkhăn\s*ướt\b',r'\bwet\s*wipes?\b',r'\bwipes?\b',
        r'\bdung\s*dịch\s*vệ\s*sinh\b',r'\blactacyd\b',r'\bgynofar\b',
        r'\bgạc\s*răng\s*miệng\b',r'\bdr\.?\s*papie\b',
    ]),
    ('Haircare & Body', 1, [
        r'\bdầu\s*gội\b',r'\bshampoo\b',
        r'\bhead\s*&?\s*shoulders?\b',r'\bpantene\b',r'\bclear\b(?=\s*(dầu|gội))',
        r'\bsunsilk\b',r'\bpalmolive\b(?=\s*(dầu|gội|silky))',
        r'\bogx\b',r'\bromano\b(?=\s*(dầu|gội|classic))',
        r'\bdầu\s*xả\b',r'\bconditioner\b',r'\bkem\s*ủ\s*tóc\b',r'\bhair\s*mask\b',
        r'\bsữa\s*tắm\b',r'\bshower\s*gel\b',r'\bbody\s*wash\b',
        r'\blifebuoy\b',r'\bdove\b(?=\s*(sữa|tắm|xà))',r'\bhazeline\b',
        r'\bvaga\b',r'\bnysha\b',r'\benchanteur\b',r'\bst\.?\s*ives\b',
        r'\bkrill\b(?=\s*(sữa|tắm))',r'\bfuji\b(?=\s*(sữa|tắm))',
        r'\bxà\s*bông\b',r'\bxà\s*phòng\b',r'\bsoap\b',
        r'\blăn\s*nách\b',r'\bdeodorant\b',r'\bkhử\s*mùi\b',r'\bnivea\b',
        r'\bnước\s*hoa\b',r'\bperfume\b',r'\bparfum\b',
        r'\baxe\b(?=\s*(nước|hoa|hương))',
        r'\bcosy\b(?=\s*(dầu|gội))',r'\bbque\b',
    ]),
    # ── PRIORITY 2: Skincare ──────────────────────────────
    ('Skincare', 2, [
        r'\bserum\b',r'\btoner\b',r'\bnước\s*hoa\s*hồng\b',r'\bessence\b',
        r'\bkem\s*dưỡng\b',r'\bkem\s*chống\s*nắng\b',r'\bsunscreen\b',r'\bspf\b',
        r'\bsữa\s*rửa\s*mặt\b',r'\bgel\s*rửa\s*mặt\b',r'\bfacial\s*(wash|cleanser)\b',
        r'\bcleanser\b',r'\bfoam\s*(rửa|wash)\b',
        r'\btẩy\s*trang\b',r'\bmicellar\b',r'\bmake[\s\-]?up\s*remover\b',
        r'\bmặt\s*nạ\b',r'\bsheet\s*mask\b',r'\bface\s*mask\b',
        r'\bkem\s*trị\s*mụn\b',r'\bmiếng\s*dán\s*mụn\b',r'\bacne\s*patch\b',
        r'\bpimple\s*patch\b',r'\bgel\s*(bôi|trị)\s*mụn\b',r'\bacmegel\b',
        r'\bneothera\b',r'\bfocallure\s*acne\b',r'\bacnes\b',
        r'\btẩy\s*da\s*chết\b',r'\bbha\b',r'\baha\b',r'\bexfoliat\b',
        r'\bsalicylic\s*acid\b',r'\bgăng\s*tay\s*tẩy\b',r'\bjary\b',
        r'\bkem\s*lót\b',r'\bkem\s*nền\b',r'\bbb\s*cream\b',r'\bcc\s*cream\b',
        r'\bcetaphil\b',r'\bcerave\b',r'\bneutrogena\b',r'\bdermablock\b',
        r'\bdermarium\b',r'\bcocoon\b',r'\bsenka\b',r'\bhadalabo\b',
        r'\banua\b',r'\bmediheal\b',r'\bnaruko\b',r'\bsur\.?\s*medic\b',
        r'\bbanobagi\b',r'\bjudydoll\b',r'\bm89\b',r'\bprobiotic\b(?=\s*(kem|dưỡng|cấp))',
        r'\bl\'oreal\b',r'\bloreal\b',r'\bst\.?\s*ives\b(?=\s*(srm|tẩy|rửa))',
        r'\bkhoáng\b(?=\s*(nước|tẩy|trang))',
        r'\bsữa\s*chống\s*nắng\b',r'\bnắng\s*spf\b',
    ]),
    # ── PRIORITY 3: Makeup ────────────────────────────────
    ('Makeup', 3, [
        r'\bson\s*(môi|tint|bóng|kem|lì|dưỡng)?\b',r'\blip\s*(stick|tint|gloss|color)\b',
        r'\bphấn\s*(má|phủ|mắt|highlight)\b',r'\bblush\b',r'\bhighlight\b',
        r'\bconceal\b',r'\bche\s*khuyết\s*điểm\b',r'\bconcealer\b',
        r'\bxịt\s*khóa\s*nền\b',r'\bsetting\s*spray\b',r'\bfixing\s*spray\b',
        r'\bkem\s*lót\b(?=\s*(trang|điểm|makeup))',
        r'\bgiấy\s*thấm\s*dầu\b',r'\boil\s*blotting\b',r'\bmerina\b',
        r'\bbông\s*(mút|tẩy)\s*(trang|điểm)\b',r'\bcọ\s*(mắt|tán|son)\b',
        r'\bdao\s*cạo\s*chân\s*mày\b',r'\bbông\s*tẩy\s*trang\b(?=\s*(calla|befou|ipek))',
        r'\bcalla\b(?=\s*(bông|tẩy))',r'\bipek\b',r'\bbefou\b',r'\buuyp\b',
        r'\bmoon\s*eyes\b',r'\bamortals\b',r'\bcolorkey\b',r'\bfoif\b',
        r'\bprany\b',r'\bmbl\b(?=\s*(xịt|khóa))',
        r'\bmascara\b',r'\beyeliner\b',r'\beyeshadow\b',
        r'\bfoundation\b',r'\bcushion\b',r'\bpowder\s*puff\b',
    ]),
]

HIGH_RISK = [
    (r'\bpantene\b(?!\s*(dầu|gội|xả))',    'Pantene: serum tóc hay dầu gội?'),
    (r'\bclear\b(?!\s*(dầu|gội))',          'Clear: dầu gội hay skincare?'),
    (r'\bdove\b(?!\s*(sữa|tắm|xà))',        'Dove: body care hay skincare?'),
    (r'\bgiấy\s*thấm\b',                    'Giấy thấm dầu: Makeup hay Skincare?'),
    (r'\bkhoáng\b',                          'Khoáng: nước tẩy trang hay nước uống?'),
    (r'\bcb\s*spf\b|\btrà\s*sữa.*spf\b',   'SPF kem chống nắng nhưng tên trà sữa?'),
    (r'\bkem\s*lót\b',                       'Kem lót: Skincare hay Makeup?'),
    (r'\bdr\.?\s*papie\b',                   'DR. Papie: dental hay baby care?'),
]

NOISE_PATTERNS = [
    (r'\btrà\s*sữa\b',          'Đồ uống xuất hiện trong Cosmetics'),
    (r'\boolong\b.*\bspf\b',    'Tên trà sữa kèm SPF — ghi nhầm category'),
    (r'^\s*\w{1,3}\s*$',        'Tên quá ngắn, không xác định'),
]

def rule_classify(text):
    if not text.strip(): return 'Khác', 0.0, ''
    tl = text.lower()
    for label, priority, patterns in sorted(RULES, key=lambda x: x[1]):
        for pat in patterns:
            try:
                if re.search(pat, tl, re.IGNORECASE):
                    conf = 1.0 if priority == 1 else (0.85 if priority == 2 else 0.7)
                    return label, conf, pat
            except: continue
    return 'Khác', 0.0, ''

def flag_risk(text):
    tl = (text or '').lower()
    for pat, reason in HIGH_RISK:
        try:
            if re.search(pat, tl, re.IGNORECASE): return reason
        except: pass
    return None

def detect_noise(text):
    tl = (text or '').lower()
    for pat, reason in NOISE_PATTERNS:
        try:
            if re.search(pat, tl, re.IGNORECASE): return reason
        except: pass
    return None

print("⏳ Gán nhãn rule-based...")
res = df['clean'].apply(rule_classify)
df['rule_label'] = [r[0] for r in res]
df['rule_conf']  = [r[1] for r in res]
df['rule_match'] = [r[2] for r in res]
df['flag_reason'] = df['item_name'].apply(flag_risk)
df['noise_flag']  = df['item_name'].apply(detect_noise)
df['flag_review'] = (df['flag_reason'].notna() | (df['rule_conf'] < 0.5)).astype(int)

print("\nPhân phối rule_label:")
print(df['rule_label'].value_counts().to_string())
n_flag = df['flag_review'].sum()
print(f"\n🚩 Cần rà soát: {n_flag} ({n_flag/len(df)*100:.1f}%)")

df_lbl = df[df['label_ground'] != 'Khác']
if len(df_lbl) > 0:
    match = (df_lbl['rule_label'] == df_lbl['label_ground']).mean()
    print(f"✅ Khớp nhãn gốc: {match*100:.1f}% trên {len(df_lbl)} mẫu")
    mis = df_lbl[df_lbl['rule_label'] != df_lbl['label_ground']][
        ['item_name','label_ground','rule_label','rule_conf']].head(20)
    print("\nTop mẫu sai nhãn:")
    print(mis.to_string(index=False))

# ─────────────────────────────────────────────────────────────
# BƯỚC 2: CHUẨN HÓA + FINAL LABEL
# ─────────────────────────────────────────────────────────────
print("\n"+"="*60); print("🔍 BƯỚC 2: CHUẨN HÓA & FINAL LABEL"); print("="*60)

# Loại bỏ noise rõ ràng (trà sữa trong cosmetics)
df_noise = df[df['noise_flag'].notna()]
print(f"⚠️  Noise phát hiện: {len(df_noise)} items")
print(df_noise[['item_name','noise_flag']].to_string(index=False))

def get_final_label(row):
    # Ground-truth LUON uu tien truoc
    if row['label_ground'] not in ['Khác', '']:
        return row['label_ground'], 'ground-truth'
    if row['noise_flag'] is not None:
        return 'Loại bỏ', 'noise'
    if row['rule_conf'] >= 0.7:
        return row['rule_label'], 'rule-based'
    return 'Cần rà soát', 'uncertain'

fl = df.apply(get_final_label, axis=1)
df['final_label']  = [r[0] for r in fl]
df['label_source'] = [r[1] for r in fl]

print("\nFinal label:")
print(df['final_label'].value_counts().to_string())
print("\nlabel_source:")
print(df['label_source'].value_counts().to_string())

# ─────────────────────────────────────────────────────────────
# BƯỚC 3: TRAIN MODEL
# ─────────────────────────────────────────────────────────────
print("\n"+"="*60); print("🤖 BƯỚC 3: TRAIN MODEL"); print("="*60)

df_tr = df[df['label_source'] == 'ground-truth'].copy()
valid = df_tr['final_label'].value_counts()
valid = valid[valid >= 3].index
df_tr = df_tr[df_tr['final_label'].isin(valid)]
print(f"Train: {len(df_tr)} rows | {df_tr['final_label'].nunique()} classes")
print(df_tr['final_label'].value_counts().to_string())

le = LabelEncoder()
y  = le.fit_transform(df_tr['final_label'])
X  = df_tr['clean'].fillna('')

# Do ít mẫu → dùng cross-validation thay vì holdout
print("\n⏳ Cross-validation (5-fold) do ít mẫu...")

# Model A: char TF-IDF
pipe_a = Pipeline([
    ('tfidf', TfidfVectorizer(ngram_range=(2,4), min_df=1, max_features=10000,
                               sublinear_tf=True, analyzer='char_wb')),
    ('clf', LogisticRegression(max_iter=500, C=3.0, class_weight='balanced'))
])
cv_a = cross_val_score(pipe_a, X, y, cv=StratifiedKFold(n_splits=min(5,min(np.bincount(y))),
                       shuffle=True, random_state=42), scoring='f1_macro')
print(f"  char TF-IDF + LR  CV F1: {cv_a.mean()*100:.2f}% ± {cv_a.std()*100:.2f}%")

# Model B: word TF-IDF
pipe_b = Pipeline([
    ('tfidf', TfidfVectorizer(ngram_range=(1,2), min_df=1, max_features=10000,
                               sublinear_tf=True)),
    ('clf', LogisticRegression(max_iter=500, C=3.0, class_weight='balanced'))
])
cv_b = cross_val_score(pipe_b, X, y, cv=StratifiedKFold(n_splits=min(5,min(np.bincount(y))),
                       shuffle=True, random_state=42), scoring='f1_macro')
print(f"  word TF-IDF + LR  CV F1: {cv_b.mean()*100:.2f}% ± {cv_b.std()*100:.2f}%")

# Train best model on full data
best_pipe = pipe_a if cv_a.mean() >= cv_b.mean() else pipe_b
best_name = 'char TF-IDF + LR' if cv_a.mean() >= cv_b.mean() else 'word TF-IDF + LR'
best_cv   = max(cv_a.mean(), cv_b.mean())
best_pipe.fit(X, y)
print(f"\n🏆 Best: {best_name}  CV F1={best_cv*100:.2f}%")

# Holdout nếu đủ mẫu
if len(df_tr) >= 20:
    X_tr, X_te, y_tr, y_te = train_test_split(X, y, test_size=0.2, random_state=42, stratify=y)
    best_pipe.fit(X_tr, y_tr)
    y_pred = best_pipe.predict(X_te)
    f1_ho = f1_score(y_te, y_pred, average='macro')
    print(f"\n📊 Holdout F1-macro: {f1_ho*100:.2f}%")
    print(classification_report(y_te, y_pred, target_names=le.classes_, zero_division=0))
    f1_per = f1_score(y_te, y_pred, average=None)
    print("Per-class F1:")
    for cls, sc in sorted(zip(le.classes_, f1_per), key=lambda x: -x[1]):
        bar = '█'*int(sc*20)+'░'*(20-int(sc*20))
        print(f"  {cls:<30} {bar} {sc*100:5.1f}%")
    # Retrain on full
    best_pipe.fit(X, y)

# ─────────────────────────────────────────────────────────────
# PREDICT ALL + EXPORT
# ─────────────────────────────────────────────────────────────
print("\n"+"="*60); print("🔮 PREDICT ALL & EXPORT"); print("="*60)

prob = best_pipe.predict_proba(df['clean'].fillna(''))
pred = np.argmax(prob, axis=1); conf = np.max(prob, axis=1)
df['ml_label'] = le.inverse_transform(pred)
df['ml_conf']  = np.round(conf, 3)

def decide(row):
    if row['label_source']=='ground-truth': return row['final_label'], 'ground-truth'
    if row['noise_flag'] is not None:      return 'Loại bỏ', 'noise'
    if row['ml_conf'] >= 0.90: return row['ml_label'], 'ml-high'
    if row['ml_conf'] >= 0.60: return row['ml_label'], 'ml-medium'
    return 'Cần rà soát', 'uncertain'

dec = df.apply(decide, axis=1)
df['pred_label']  = [r[0] for r in dec]
df['pred_source'] = [r[1] for r in dec]

print("✅ Pred label:"); print(df['pred_label'].value_counts().to_string())

cols = ['category','user_id','product_group','item_name','clean',
        'label_ground','rule_label','rule_conf','rule_match',
        'ml_label','ml_conf','pred_label','pred_source',
        'flag_review','flag_reason','noise_flag']

df[cols].to_csv(OUT+'cosmetics_full_labeled.csv', index=False, encoding='utf-8-sig')
print(f"\n✅ cosmetics_full_labeled.csv ({len(df)} rows)")

df_rft = df[df['pred_source'].isin(['ground-truth','ml-high']) & (df['pred_label']!='Loại bỏ')]
df_rft[['item_name','clean','pred_label']].to_csv(OUT+'cosmetics_ready_for_train.csv', index=False, encoding='utf-8-sig')
print(f"✅ cosmetics_ready_for_train.csv ({len(df_rft)} rows)")

df_rev = df[df['pred_source'].isin(['uncertain','ml-medium'])][cols]
df_rev.to_csv(OUT+'cosmetics_review_queue.csv', index=False, encoding='utf-8-sig')
print(f"✅ cosmetics_review_queue.csv ({len(df_rev)} rows)")

rules_rows = []
for label, prio, patterns in RULES:
    for p in patterns:
        rules_rows.append({'label':label,'priority':prio,'pattern':p})
pd.DataFrame(rules_rows).to_csv(OUT+'cosmetics_keyword_rules_v1.csv', index=False, encoding='utf-8-sig')
print(f"✅ cosmetics_keyword_rules_v1.csv ({len(rules_rows)} rules)")

# Noise report
if len(df_noise) > 0:
    df_noise[['item_name','product_group','noise_flag']].to_csv(OUT+'cosmetics_noise_report.csv', index=False, encoding='utf-8-sig')
    print(f"✅ cosmetics_noise_report.csv ({len(df_noise)} rows)")

summary = {
    'total_items': int(len(df)),
    'sub_groups': list(le.classes_),
    'cv_f1_char_lr': f"{cv_a.mean()*100:.2f}%",
    'cv_f1_word_lr': f"{cv_b.mean()*100:.2f}%",
    'best_model': best_name,
    'best_cv_f1': f"{best_cv*100:.2f}%",
    'ready_for_train': int(len(df_rft)),
    'review_queue': int(len(df_rev)),
    'noise_removed': int(len(df_noise)),
    'keyword_rules': len(rules_rows),
}
with open(OUT+'cosmetics_summary.json','w',encoding='utf-8') as f:
    json.dump(summary, f, ensure_ascii=False, indent=2)

print(f"""
╔══════════════════════════════════════════════════════╗
║        COSMETICS PIPELINE — SUMMARY                 ║
╠══════════════════════════════════════════════════════╣
║  Tổng items            : {len(df):>6}                    ║
║  CV F1 char TF-IDF+LR  : {cv_a.mean()*100:>6.2f}%                  ║
║  CV F1 word TF-IDF+LR  : {cv_b.mean()*100:>6.2f}%                  ║
║  Best model            : {best_name:<26}  ║
║  Best CV F1            : {best_cv*100:>6.2f}%                  ║
║  Ready-for-train       : {len(df_rft):>6}                    ║
║  Review queue          : {len(df_rev):>6}                    ║
║  Noise removed         : {len(df_noise):>6}                    ║
║  Keyword rules         : {len(rules_rows):>6}                    ║
╚══════════════════════════════════════════════════════╝
""")