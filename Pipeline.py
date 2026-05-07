"""
Pipeline V2 — fix label inconsistency + improve model
"""
import pandas as pd, numpy as np, re, unicodedata, warnings
warnings.filterwarnings('ignore')

from openpyxl import load_workbook
from underthesea import word_tokenize
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report, f1_score
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import LabelEncoder
import xgboost as xgb

# ── Load ────────────────────────────────────────────────────
print("📂 Loading data…")
DATA_PATH = 'Data_2025_2026_Combined_Cleaned.xlsx'
wb = load_workbook(DATA_PATH, read_only=True)
COLS = ['category_original','user_id','product_group','item_name']
dfs = []
for sheet in wb.sheetnames:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2: continue
    clean = [r[:4] for r in rows[1:] if len(r)>=4 and any(v is not None for v in r[:4])]
    if not clean: continue
    df = pd.DataFrame(clean, columns=COLS)
    df['sheet'] = sheet
    dfs.append(df)
df_all = pd.concat(dfs, ignore_index=True)
for c in COLS:
    df_all[c] = df_all[c].fillna('').astype(str).str.strip()
print(f"  {len(df_all):,} rows, {df_all['sheet'].nunique()} sheets")

# ── Chuẩn hóa nhãn gốc ──────────────────────────────────────
LABEL_NORM = {
    'Cosmetic':       'Cosmetics',
    'Accesories':     'Accessories',
    'Dịch vụ':        'Service',
    'Pets':           'Pet',
    'Beverages':      'Beverages',
    'Food':           'Food',
    'Medical':        'Medical',
    'Clothing':       'Clothing',
    'Cosmetics':      'Cosmetics',
    'Books':          'Books',
    'Home & Office':  'Home & Office',
    'Service':        'Service',
    'Electronics':    'Electronics',
    'Accessories':    'Accessories',
    'Jewelry':        'Jewelry',
    'Entertainment':  'Entertainment',
    'Pet':            'Pet',
}
df_all['label_norm'] = df_all['category_original'].map(LABEL_NORM).fillna('Other')
print("\n📊 Nhãn sau chuẩn hóa:")
print(df_all['label_norm'].value_counts().to_string())

# ── Preprocessing ───────────────────────────────────────────
TYPO_MAP = {
    r'\bchan ga\b':'chân gà', r'\bnem gion\b':'nem giòn',
    r'\bxuc xich\b':'xúc xích', r'\bthit bo\b':'thịt bò',
    r'\bthit heo\b':'thịt heo', r'\bca loc\b':'cá lóc',
    r'\bca hoi\b':'cá hồi', r'\bdau hao\b':'dầu hào',
    r'\bnuoc mam\b':'nước mắm', r'\bbanh mi\b':'bánh mì',
    r'\brau muong\b':'rau muống', r'\bxoai\b':'xoài',
    r'\bdua leo\b':'dưa leo',
}
UNIT_RE = re.compile(
    r'\b\d+[\.,]?\d*\s*(?:kg|g|gram|ml|l|lít|lon|chai|cái|hộp|túi|pack|'
    r'gói|thùng|bịch|thanh|miếng|lát|trái|quả|củ|bó|set|pcs?|pieces?)\b',
    re.IGNORECASE)

def preprocess(text):
    if not text or not text.strip(): return ''
    t = unicodedata.normalize('NFC', text).lower()
    for p, r in TYPO_MAP.items():
        t = re.sub(p, r, t, flags=re.IGNORECASE)
    t = UNIT_RE.sub(' ', t)
    t = re.sub(r'[^\w\s\u00C0-\u024F\u1E00-\u1EFF\-]', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    try: t = word_tokenize(t, format='text')
    except: pass
    return t

print("\n⏳ Preprocessing…")
df_all['clean'] = df_all['item_name'].apply(preprocess)
print("  Done.")

# ── Rule-based ───────────────────────────────────────────────
CATEGORY_RULES = {
    'Beverages_Alcoholic': ['bia','beer','rượu','wine','vang','whisky','whiskey','vodka','gin','rum','brandy','cognac','sake','soju','strongbow','heineken','tiger','budweiser','corona','cider','champagne','prosecco','cồn'],
    'Beverages_NonAlcoholic': ['nước ngọt','coca','pepsi','sprite','7up','fanta','nước tăng lực','red bull','sting','monster','nước suối','nước khoáng','trà','tea','matcha','hồng trà','trà xanh','cà phê','cafe','coffee','latte','espresso','sinh tố','smoothie','nước ép','juice','sữa tươi','sữa đậu nành','trà sữa','bubble tea','boba','milk tea','kombucha','yakult','yogurt uống','nước dừa'],
    'Food_Meat_Seafood': ['thịt bò','thịt heo','thịt lợn','thịt gà','gà','chân gà','thịt xay','thịt nạc','thịt ba chỉ','sườn','xương','ếch','chim cút','vịt','ngan','cá','cá lóc','cá hồi','cá thu','cá ngừ','tôm','mực','cua','ghẹ','nghêu','sò','sashimi','hải sản','seafood','meat deli'],
    'Food_Fresh_Produce': ['rau','cải','bí đao','bí đỏ','bắp cải','rau muống','cải thìa','xà lách','khoai lang','khoai tây','khoai môn','củ cải','xoài','ổi','táo','lê','cam','quýt','bưởi','dưa leo','dưa hấu','mít','vải','nhãn','chuối','dâu','nho','thanh long','đu đủ','bắp','ngô','đậu','cà chua','cà rốt','hành','tỏi','gừng','sả','nấm','măng','dứa','khổ qua'],
    'Food_Processed_Ready': ['cơm trộn','cơm tấm','cơm rang','gà chiên','gà nướng','pizza','burger','sandwich','bánh mì kẹp','phở','bún','hủ tiếu','miến','mì','lẩu','kimbap','gimbap','sushi','nem nướng','chả giò','bánh cuốn','bánh xèo','dimsum','há cảo','xíu mại','bánh bao','bò né','bít tết','steak','pasta'],
    'Food_Snack_Confectionery': ['snack','oishi','bánh quy','kẹo','chocolate','bim bim','cookies','richy','đậu phộng','bouchee','lotte','anytime','kokola','jojo','hạt','popcorn','chips','wafer','cracker','oreo','kitkat','milo','kẹo dẻo','kẹo bạc hà','socola','thạch'],
    'Food_Dairy_Eating': ['sữa chua ăn','yogurt ăn','phô mai','cheese','bơ','butter','kem que','kem','ice cream','bánh flan','pudding'],
    'Food_Dry_Condiment': ['xúc xích','pate','nem chua','chả lụa','giò lụa','thịt nguội','lạp xưởng','dầu hào','nước mắm','dầu ăn','giấm','bột nêm','gia vị','muối','đường','tương ớt','tương đen','maggi','knorr','mì gói','acecook','hảo hảo','omachi','kokomi','bánh pía','bánh tét','nước sốt','mayonnaise'],
    'Food_Starchy_Grain': ['gạo','rice','bột mì','flour','ngũ cốc','granola','yến mạch','oat'],
    'Clothing': ['áo','quần','váy','đầm','đồ bộ','áo sơ mi','áo thun','áo hoodie','áo khoác','áo len','quần jean','quần short','quần tây','đồ lót','nội y','underwear','bra','boxer','giày','dép','sandal','sneaker','boots','mũ','nón','tất','vớ'],
    'Cosmetics_Skincare': ['toner','serum','kem dưỡng','moisturizer','lotion','kem chống nắng','sunscreen','spf','tẩy trang','sữa rửa mặt','cleanser','mặt nạ','mask','essence','retinol','niacinamide','hyaluronic'],
    'Cosmetics_Makeup': ['son','lipstick','lip','phấn','foundation','cushion','bb cream','cc cream','mascara','eyeliner','eyeshadow','blush','highlight','concealer','kẻ mắt','má hồng'],
    'Cosmetics_Haircare': ['dầu gội','dầu xả','shampoo','conditioner','sáp vuốt tóc','wax tóc','hair mask','kem ủ tóc','xịt tóc'],
    'Cosmetics_Personal_Care': ['sữa tắm','shower gel','xà phòng','soap','nước rửa tay','hand wash','kem đánh răng','toothpaste','bàn chải','toothbrush','dao cạo','razor','lăn khử mùi','deodorant','nước hoa','perfume','băng vệ sinh','tampon','bỉm','tã'],
    'Accessories': ['túi','balo','ví','wallet','bag','purse','bình nước','water bottle','thermos','băng đô','chun tóc','kẹp tóc','găng tay','gloves','gọng kính','kính mắt','dây đeo','strap','lót giày','khăn','towel','ô dù'],
    'Electronics': ['tivi','tv','television','tủ lạnh','refrigerator','máy giặt','máy lạnh','điều hòa','laptop','macbook','máy tính','điện thoại','smartphone','iphone','samsung','oppo','xiaomi','tai nghe','earphone','headphone','airpods','màn hình','monitor','keyboard','chuột','bếp từ','lò vi sóng','microwave','nồi cơm điện','quạt','máy hút bụi','sạc','charger','cáp','cable'],
    'Home_Office': ['nước giặt','nước rửa chén','nước lau sàn','chất tẩy','bleach','vim','đèn','bóng đèn','lamp','hộp lưu trữ','kệ','khăn trải bàn','gối','pillow','chăn','băng keo','kẹp giấy','bút','vở','máy in','mực in','sáp thơm','nến','candle','tinh dầu','chổi','cây lau','khăn lau'],
    'Medical': ['thuốc','medicine','paracetamol','ibuprofen','aspirin','vitamin','thực phẩm chức năng','supplement','collagen','canxi','calcium','omega','probiotics','băng','bông','gạc','antiseptic','khẩu trang','nhiệt kế','huyết áp','dung dịch vệ sinh'],
    'Pet': ['thức ăn chó','thức ăn mèo','dog food','cat food','hạt cho chó','hạt cho mèo','pate chó','pate mèo','vòng cổ chó','dây dắt','leash','collar','cát vệ sinh mèo','cat litter','pet toy','lồng','cage','vitamin thú','thuốc rận','royal canin','whiskas','pedigree','hills'],
    'Books': ['sách','book','tạp chí','magazine','báo','newspaper','truyện','manga','comic','sách giáo khoa','textbook','giáo trình'],
    'Service': ['spa','massage','mát xa','gym','fitness','vé','ticket','vé máy bay','vé tàu','vé xe','hoa','flower','bouquet','ngân hàng','banking','bảo hiểm','insurance'],
    'Entertainment': ['game','gaming','karaoke','cinema','phim','movie','đồ chơi','toy','lego','puzzle','board game','sở thú','zoo','concert'],
    'Jewelry': ['nhẫn','ring','bông tai','earring','dây chuyền','necklace','mặt dây','pendant','vòng tay','bracelet','đồng hồ','watch','kim cương','diamond','vàng','gold','bạc','silver'],
}
GROUP_MAP = {
    'Beverages_Alcoholic':'Beverages','Beverages_NonAlcoholic':'Beverages',
    'Food_Meat_Seafood':'Food','Food_Fresh_Produce':'Food','Food_Processed_Ready':'Food',
    'Food_Snack_Confectionery':'Food','Food_Dairy_Eating':'Food',
    'Food_Dry_Condiment':'Food','Food_Starchy_Grain':'Food',
    'Clothing':'Clothing',
    'Cosmetics_Skincare':'Cosmetics','Cosmetics_Makeup':'Cosmetics',
    'Cosmetics_Haircare':'Cosmetics','Cosmetics_Personal_Care':'Cosmetics',
    'Accessories':'Accessories','Electronics':'Electronics',
    'Home_Office':'Home & Office','Medical':'Medical',
    'Pet':'Pet','Books':'Books','Service':'Service',
    'Entertainment':'Entertainment','Jewelry':'Jewelry',
}
HIGH_RISK = [
    (r'yogurt|sữa chua',               'Yogurt có thể ăn hoặc uống'),
    (r'cookies|cream|freeze|đá xay',   'Topping kem hoặc đồ uống xay?'),
    (r'strongbow|cider|kombucha',       'Lẫn có cồn/không cồn'),
    (r'nem|bánh mì bò',                'Món ăn sẵn hay nguyên liệu?'),
    (r'nước ép|juice|sinh tố',         'Rau quả tươi vs đồ uống'),
    (r'vitamin|supplement|thực phẩm chức năng', 'Dược phẩm vs thực phẩm'),
    (r'phô mai|cheese|bơ',             'Dairy ăn vs nguyên liệu'),
    (r'đồ chơi|toy',                    'Trẻ em vs thú cưng'),
    (r'băng vệ sinh|tampon',           'Medical vs Personal Care'),
]

def rule_classify(text):
    if not text: return 'Other','Other',0.0
    tl = text.lower()
    scores = {l: sum(1 for kw in kws if kw.lower() in tl) for l,kws in CATEGORY_RULES.items()}
    scores = {l:s for l,s in scores.items() if s>0}
    if not scores: return 'Other','Other',0.0
    best = max(scores, key=scores.get)
    conf = min(scores[best]/3.0, 1.0)
    return best, GROUP_MAP.get(best,'Other'), round(conf,2)

def flag_risk(text):
    if not text: return None
    for pat,reason in HIGH_RISK:
        if re.search(pat, text.lower(), re.IGNORECASE): return reason
    return None

print("⏳ Rule-based…")
res = df_all['clean'].apply(rule_classify)
df_all['rule_sub']   = [r[0] for r in res]
df_all['rule_group'] = [r[1] for r in res]
df_all['rule_conf']  = [r[2] for r in res]

df_all['flag_reason'] = df_all['clean'].apply(flag_risk)
df_all['flag_review'] = (
    df_all['flag_reason'].notna() |
    (df_all['rule_conf'] < 0.3) |
    (df_all['rule_sub'] == 'Other')
).astype(int)

n_flag = df_all['flag_review'].sum()
print(f"  Cần rà soát: {n_flag:,} ({n_flag/len(df_all)*100:.1f}%)")

# ── ML — dùng label_norm đã chuẩn hóa ──────────────────────
print("\n🤖 ML Pipeline…")
df_tr = df_all[df_all['label_norm'] != 'Other'].copy()
valid  = df_tr['label_norm'].value_counts()
valid  = valid[valid >= 5].index
df_tr  = df_tr[df_tr['label_norm'].isin(valid)]
print(f"  Train size: {len(df_tr):,}, Classes: {df_tr['label_norm'].nunique()}")
print(df_tr['label_norm'].value_counts().to_string())

le = LabelEncoder()
y  = le.fit_transform(df_tr['label_norm'])
X  = df_tr['clean'].fillna('')
X_tr, X_te, y_tr, y_te = train_test_split(X, y, test_size=0.2, random_state=42, stratify=y)

# LR
print("\n⏳ Logistic Regression…")
pipe_lr = Pipeline([
    ('tfidf', TfidfVectorizer(ngram_range=(1,2), min_df=2, max_features=30000, sublinear_tf=True)),
    ('clf',   LogisticRegression(max_iter=1000, C=5.0, class_weight='balanced'))
])
pipe_lr.fit(X_tr, y_tr)
y_lr = pipe_lr.predict(X_te)
f1_lr = f1_score(y_te, y_lr, average='macro')
print(f"  LR  F1-macro: {f1_lr*100:.2f}%")

# XGB
print("⏳ XGBoost…")
tfidf = TfidfVectorizer(ngram_range=(1,2), min_df=2, max_features=30000, sublinear_tf=True)
Xtr_t = tfidf.fit_transform(X_tr)
Xte_t = tfidf.transform(X_te)
xgb_m = xgb.XGBClassifier(n_estimators=300, max_depth=6, learning_rate=0.1,
                            eval_metric='mlogloss', random_state=42, n_jobs=-1)
xgb_m.fit(Xtr_t, y_tr)
y_xgb = xgb_m.predict(Xte_t)
f1_xgb = f1_score(y_te, y_xgb, average='macro')
print(f"  XGB F1-macro: {f1_xgb*100:.2f}%")

use_xgb  = f1_xgb > f1_lr
best_name= 'XGBoost' if use_xgb else 'LogisticRegression'
best_f1  = max(f1_lr, f1_xgb)
y_best   = y_xgb if use_xgb else y_lr
print(f"\n🏆 Best: {best_name}  F1={best_f1*100:.2f}%")
print("\n📊 Classification Report:")
print(classification_report(y_te, y_best, target_names=le.classes_, zero_division=0))

# Predict all
Xall = df_all['clean'].fillna('')
if use_xgb:
    prob = xgb_m.predict_proba(tfidf.transform(Xall))
else:
    prob = pipe_lr.predict_proba(Xall)
pred = np.argmax(prob, axis=1)
conf = np.max(prob, axis=1)
df_all['ml_label'] = le.inverse_transform(pred)
df_all['ml_conf']  = np.round(conf, 3)

def final_label(row):
    if row['flag_review']==0 and row['rule_conf']>=0.5:
        return row['rule_group'], 'rule-based'
    if row['ml_conf']>=0.85: return row['ml_label'], 'ml-high'
    if row['ml_conf']>=0.5:  return row['ml_label'], 'ml-medium'
    return 'Cần rà soát', 'uncertain'

fl = df_all.apply(final_label, axis=1)
df_all['final_label']  = [r[0] for r in fl]
df_all['label_source'] = [r[1] for r in fl]

print("\n✅ Final label distribution:")
print(df_all['final_label'].value_counts().to_string())

# ── Export ───────────────────────────────────────────────────
OUT = 'EDA_Outputs/'
cols_out = ['sheet','category_original','label_norm','user_id','product_group','item_name','clean',
            'rule_sub','rule_group','rule_conf','ml_label','ml_conf',
            'final_label','label_source','flag_review','flag_reason']
df_all[cols_out].to_csv(OUT+'full_classified_v2.csv', index=False, encoding='utf-8-sig')

df_review = df_all[df_all['flag_review']==1][cols_out].sort_values(['flag_reason','rule_conf'])
df_review.to_csv(OUT+'review_queue_v2.csv', index=False, encoding='utf-8-sig')

sub_sum = df_all.groupby(['final_label','rule_sub']).size().reset_index(name='count').sort_values(['final_label','count'],ascending=[True,False])
sub_sum.to_csv(OUT+'category_sublabel_summary_v2.csv', index=False, encoding='utf-8-sig')

# Food deep-dive
df_food = df_all[df_all['final_label']=='Food'].copy()
food_sub = df_food['rule_sub'].value_counts().reset_index()
food_sub.columns = ['sub_group','count']
food_sub.to_csv(OUT+'food_subgroup_breakdown.csv', index=False, encoding='utf-8-sig')

print(f"""
╔══════════════════════════════════════════════════╗
║           PIPELINE V2 — SUMMARY                 ║
╠══════════════════════════════════════════════════╣
║  Tổng item         : {len(df_all):>10,}            ║
║  LR  F1-macro      : {f1_lr*100:>9.2f}%            ║
║  XGB F1-macro      : {f1_xgb*100:>9.2f}%            ║
║  Best model        : {best_name:<24}║
║  Best F1-macro     : {best_f1*100:>9.2f}%            ║
║  Cần rà soát       : {n_flag:>10,}            ║
╚══════════════════════════════════════════════════╝
""")