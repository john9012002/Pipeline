"""
================================================================
FOOD CLASSIFICATION PIPELINE — PHA 2
Bước 1: Rule-based labeling (8 sub-groups)
Bước 2: Chuẩn hóa nhãn + review list
Bước 3: Train model + predict + export
================================================================
"""
import pandas as pd, numpy as np, re, unicodedata, warnings, json
warnings.filterwarnings('ignore')
from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import f1_score, classification_report
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import LabelEncoder
import xgboost as xgb

OUT = 'D:\\EDA\\Category_pipelines\\food'

# ─────────────────────────────────────────────────────────────
# 0. LOAD DATA
# ─────────────────────────────────────────────────────────────
print("="*60); print("📂 0. LOAD DATA"); print("="*60)
wb = load_workbook('Data_2025_2026_Combined_Cleaned.xlsx', read_only=True)
ws = wb['Food']
rows = list(ws.iter_rows(values_only=True))
data = [r[:4] for r in rows[1:] if r and any(v is not None for v in r[:4])]
df = pd.DataFrame(data, columns=['category','user_id','product_group','item_name'])
df['item_name']     = df['item_name'].fillna('').astype(str).str.strip()
df['product_group'] = df['product_group'].fillna('').astype(str).str.strip()
print(f"✅ {len(df):,} rows | {df['product_group'].nunique()} product groups")

# ─────────────────────────────────────────────────────────────
# 0b. CHUẨN HÓA NHÃN GỐC → 8 sub-groups
# ─────────────────────────────────────────────────────────────
GROUP_NORM = {
    # Thịt - Hải sản - Trứng
    'Thịt Heo':'Thịt - Hải sản - Trứng','Thịt Gà':'Thịt - Hải sản - Trứng',
    'Thịt Bò':'Thịt - Hải sản - Trứng','Thịt Cá':'Thịt - Hải sản - Trứng',
    'Thịt Tôm':'Thịt - Hải sản - Trứng',
    'Cá':'Thịt - Hải sản - Trứng','Tôm':'Thịt - Hải sản - Trứng',
    'Mực':'Thịt - Hải sản - Trứng','Cua':'Thịt - Hải sản - Trứng',
    'Hải Sản':'Thịt - Hải sản - Trứng',
    'Trứng gà':'Thịt - Hải sản - Trứng','Trứng vịt':'Thịt - Hải sản - Trứng',
    # Rau - Củ - Quả
    'Rau':'Rau - Củ - Quả','Trái Cây':'Rau - Củ - Quả',
    'Chuối':'Rau - Củ - Quả','Nho':'Rau - Củ - Quả',
    'Xoài':'Rau - Củ - Quả','Dâu Tây':'Rau - Củ - Quả',
    'Cà Chua':'Rau - Củ - Quả','Nấm':'Rau - Củ - Quả',
    'Dây tây':'Rau - Củ - Quả',
    # Đồ ăn chế biến sẵn
    'Bún':'Đồ ăn chế biến sẵn','Phở':'Đồ ăn chế biến sẵn',
    'Cơm chiên':'Đồ ăn chế biến sẵn','Cháo':'Đồ ăn chế biến sẵn',
    'Bánh mì':'Đồ ăn chế biến sẵn','Pizza':'Đồ ăn chế biến sẵn',
    'Chả Giò':'Đồ ăn chế biến sẵn','Xôi':'Đồ ăn chế biến sẵn',
    'Lẩu':'Đồ ăn chế biến sẵn','Bánh bao':'Đồ ăn chế biến sẵn',
    'Bánh Tráng Trộn':'Đồ ăn chế biến sẵn','Kim Chi':'Đồ ăn chế biến sẵn',
    'Suất Ăn':'Đồ ăn chế biến sẵn','Combo Food':'Đồ ăn chế biến sẵn',
    'Buffet':'Đồ ăn chế biến sẵn','Xuất Ăn':'Đồ ăn chế biến sẵn',
    # Snack - Bánh kẹo
    'Snack':'Snack - Bánh kẹo','Bánh Quy':'Snack - Bánh kẹo',
    'Bánh Tráng':'Snack - Bánh kẹo',
    # Mì - Ngũ cốc khô
    'Mì Gói':'Mì - Ngũ cốc - Khô','Mì gói':'Mì - Ngũ cốc - Khô',
    # Gia vị - Đồ hộp - Chế biến
    'Gia Vị':'Gia vị - Sốt - Dầu','Tương Ớt':'Gia vị - Sốt - Dầu',
    'Nước Mắm':'Gia vị - Sốt - Dầu','Mật Ong':'Gia vị - Sốt - Dầu',
    'Dầu Ăn':'Gia vị - Sốt - Dầu','Muối':'Gia vị - Sốt - Dầu',
    'Xúc xích':'Gia vị - Sốt - Dầu',
    # Dairy - Trứng ăn
    'Kem lạnh':'Dairy - Kem - Phô mai','Phô Mai':'Dairy - Kem - Phô mai',
    'Yến Chưng':'Dairy - Kem - Phô mai',
}
df['label_ground'] = df['product_group'].map(GROUP_NORM).fillna('Khác')
print("\nNhãn sau chuẩn hóa:")
print(df['label_ground'].value_counts().to_string())

# ─────────────────────────────────────────────────────────────
# 1. PREPROCESSING
# ─────────────────────────────────────────────────────────────
print("\n"+"="*60); print("⚙️  1. PREPROCESSING"); print("="*60)
TYPO = {
    r'\bthit\b':'thịt',r'\bga\b':'gà',r'\bheo\b':'heo',r'\bbo\b':'bò',
    r'\bca\b':'cá',r'\btom\b':'tôm',r'\bmuc\b':'mực',r'\bcua\b':'cua',
    r'\brau\b':'rau',r'\btrung\b':'trứng',r'\bbanh\b':'bánh',
    r'\bmi\b':'mì',r'\bpho\b':'phở',r'\bbun\b':'bún',r'\bchao\b':'cháo',
    r'\bcom\b':'cơm',r'\bsnack\b':'snack',r'\bkem\b':'kem',
    r'\bgia vi\b':'gia vị',r'\bnuoc mam\b':'nước mắm',
    r'\bdau an\b':'dầu ăn',r'\btuong ot\b':'tương ớt',
    r'\bxuc xich\b':'xúc xích',r'\bpho mai\b':'phô mai',
}
UNIT_RE = re.compile(
    r'\b\d+[\.,]?\d*\s*(?:kg|g|gram|ml|l|gói|hộp|túi|cái|miếng|'
    r'lát|trái|quả|củ|bó|lon|chai|thùng|bịch|set|pcs?|phần|suất|vé)\b',
    re.IGNORECASE)
SIZE_RE = re.compile(r'\b(size\s*)?[smlxl]+\b|\b(small|medium|large)\b', re.IGNORECASE)

def preprocess(text):
    if not text or not text.strip(): return ''
    t = unicodedata.normalize('NFC', text).lower()
    for p, r in TYPO.items():
        t = re.sub(p, r, t, flags=re.IGNORECASE)
    t = UNIT_RE.sub(' ', t)
    t = SIZE_RE.sub(' ', t)
    t = re.sub(r'[^\w\s\u00C0-\u024F\u1E00-\u1EFF\-&+/]', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t

df['clean'] = df['item_name'].apply(preprocess)
print(f"✅ Preprocessing xong {len(df):,} items")

# ─────────────────────────────────────────────────────────────
# BƯỚC 1: RULE-BASED LABELING
# ─────────────────────────────────────────────────────────────
print("\n"+"="*60); print("📏 BƯỚC 1: RULE-BASED LABELING"); print("="*60)

RULES = [
    # Priority 1 — rõ nhất
    ('Thịt - Hải sản - Trứng', 1, [
        r'\bthịt\s*(heo|lợn|bò|gà|vịt|dê|ngan|bồ câu|thỏ)\b',
        r'\bthịt\s*xay\b',r'\bthịt\s*ba\s*chỉ\b',r'\bsườn\b',r'\bxương\b',
        r'\bba\s*rọi\b',r'\bnạc\s*vai\b',r'\bnạc\s*đùi\b',r'\bcốt\s*lết\b',
        r'\bchân\s*gà\b',r'\bgà\s*(nướng|chiên|luộc|rang|hấp)\b',
        r'\bcá\s*(nục|hồi|thu|basa|ngừ|tuyết|lóc|tra|bống|hake)\b',
        r'\btôm\b',r'\bmực\b',r'\bcua\b',r'\bghẹ\b',r'\bnghêu\b',
        r'\bhào\b',r'\bsò\b',r'\btrứng\s*(gà|vịt|cút|omega)\b',
        r'\bhải\s*sản\b',r'\bseafood\b',r'\bsashimi\b',
        r'\bmeat\s*deli\b',r'\bc\.p\b',r'\bvissan\b',
    ]),
    ('Rau - Củ - Quả', 1, [
        r'\brau\s*(cải|muống|mồng tơi|dền|ngót|lang|nhút|má|bina|bó xôi)\b',
        r'\bcải\s*(thảo|xanh|thìa|ngọt|bắp|trắng)\b',r'\bbắp\s*cải\b',
        r'\bxà\s*lách\b',r'\bsalad\s*(leaves?|xanh)?\b',
        r'\bnấm\s*(kim châm|rơm|đông cô|linh chi|bào ngư|hương)\b',
        r'\bkhoai\s*(tây|lang|môn|mỡ)\b',r'\bcủ\s*cải\b',r'\bcà\s*rốt\b',
        r'\bhành\s*(tây|lá|tím)\b',r'\btỏi\b',r'\bgừng\b',r'\bsả\b',
        r'\bớt\s*(tươi|đỏ|xanh)\b',r'\bbắp\b',r'\bngô\b',r'\bđậu\b',
        r'\bcà\s*chua\b',r'\bdưa\s*(leo|chuột|hấu|lưới)\b',
        r'\bkh[oố]\s*qua\b',r'\bbí\s*(đao|đỏ|ngô)\b',r'\bmăng\b',
        r'\bxoài\b',r'\bổi\b',r'\btáo\b',r'\blê\b',r'\bcam\b',r'\bquýt\b',
        r'\bbưởi\b',r'\bchuối\b',r'\bdâu\b',r'\bnho\b',r'\bthanh\s*long\b',
        r'\bđu\s*đủ\b',r'\bchôm\s*chôm\b',r'\bmít\b',r'\bvải\b',r'\bnhãn\b',
        r'\bdứa\b',r'\btrái\s*cây\b',r'\bfruit\b',
    ]),
    # Priority 2
    ('Snack - Bánh kẹo', 2, [
        r'\bsnack\b',r'\boishi\b',r'\bbánh\s*quy\b',r'\bcookies?\b',
        r'\bchips?\b',r'\bpopcorn\b',r'\bbỏng\s*ngô\b',
        r'\bbánh\s*tráng\b',r'\bwafer\b',r'\bcracker\b',
        r'\bkẹo\b',r'\bchocolate\b',r'\bsocola\b',r'\bnougat\b',
        r'\bthạch\b',r'\bjelly\b',r'\bhạt\s*(điều|bí|hướng dương|macca)\b',
        r'\bđậu\s*phộng\b',r'\bpeanut\b',r'\blotte\b',r'\brichy\b',
        r'\boreo\b',r'\bkitkat\b',r'\bbouchee\b',r'\bpocky\b',
        r'\bpillows?\b',r'\bsnack\s*bắp\b',r'\bbánh\s*phồng\b',
    ]),
    ('Mì - Ngũ cốc - Khô', 2, [
        r'\bmì\s*gói\b',r'\bmì\s*tôm\b',r'\bmì\s*xào\b',r'\bmì\s*trộn\b',
        r'\bacecook\b',r'\bhảo\s*hảo\b',r'\bomachi\b',r'\bkokomi\b',
        r'\bvifon\b',r'\bgoreng\b',r'\bmaggi\b',
        r'\bgạo\b',r'\bcơm\s*(tấm|trắng)?\b(?!chiên|rang|xào)',
        r'\bnugg?ets?\b',r'\bbột\s*mì\b',r'\bngũ\s*cốc\b',
        r'\bgranola\b',r'\byến\s*mạch\b',r'\boat\b',r'\bcornflakes\b',
        r'\bmiến\b',r'\bhủ\s*tiếu\b',
    ]),
    # Priority 3
    ('Gia vị - Sốt - Dầu', 3, [
        r'\bnước\s*mắm\b',r'\bdầu\s*ăn\b',r'\bdầu\s*hào\b',r'\bdầu\s*mè\b',
        r'\bgiấm\b',r'\bbột\s*nêm\b',r'\bgía\s*vị\b',r'\bgia\s*vị\b',
        r'\bmuối\b',
        r'\bđường\b',r'\btiêu\b',r'\bột\s*ớt\b',r'\bbột\s*tỏi\b',
        r'\btương\s*(ớt|đen|hoisin|barbecue)\b',r'\bchinsu\b',
        r'\bmaggi\s*sốt\b',r'\bknorr\b',r'\bmasan\b',
        r'\bnước\s*sốt\b',r'\bsốt\b',r'\bmayonnaise\b',r'\bmù\s*tạt\b',
        r'\bketchup\b',r'\bbơ\b',
        r'\bmật\s*ong\b',r'\bhoney\b',
        r'\bdầu\s*oliu\b',r'\bolive\s*oil\b',
        r'\bxúc\s*xích\b',r'\bpate\b',r'\bgiò\s*lụa\b',r'\bchả\s*lụa\b',
        r'\bnem\s*chua\b',r'\blạp\s*xưởng\b',r'\bthịt\s*nguội\b',
    ]),
    ('Dairy - Kem - Phô mai', 3, [
        r'\bkem\s*(que|hộp|ly|cone|lạnh|socola|vani|trà sữa)\b',
        r'\bice\s*cream\b',r'\bcelano\b',r'\bwalls?\b',r'\bmerino\b',
        r'\bphô\s*mai\b',r'\bcheese\b',r'\bmozzarella\b',
        r'\bpudding\b',r'\bflan\b',r'\bmousse\b',
        r'\byến\s*chưng\b',r'\bnước\s*yến\b',r'\byến\s*sào\b',
    ]),
    # Priority 4
    ('Đồ ăn chế biến sẵn', 4, [
        r'\bbún\b',r'\bphở\b',r'\bhủ\s*tiếu\b',r'\bbánh\s*canh\b',
        r'\bcháo\b',r'\bxôi\b',r'\bcơm\s*(chiên|rang|tấm|trộn)\b',
        r'\bgà\s*(chiên|nướng|rang muối|xào sả ớt)\b',
        r'\bbánh\s*mì\b',r'\bsandwich\b',r'\bburger\b',r'\bpizza\b',
        r'\bpasta\b',r'\bmì\s*ý\b',r'\bspaghetti\b',
        r'\bkimbap\b',r'\bgimbap\b',r'\bsushi\b',r'\bsashimi\b',
        r'\bdimsum\b',r'\bhá\s*cảo\b',r'\bxíu\s*mại\b',r'\bbánh\s*bao\b',
        r'\bchả\s*giò\b',r'\bnem\s*nướng\b',r'\bbánh\s*cuốn\b',
        r'\bbánh\s*xèo\b',r'\bsteak\b',r'\bbít\s*tết\b',r'\bbò\s*né\b',
        r'\blẩu\b',r'\bhotpot\b',r'\bkim\s*chi\b',r'\bkimchi\b',
        r'\bbuffet\b',r'\bsuất\s*ăn\b',r'\bcombo\b',r'\bset\s*(ăn|cơm|lẩu)\b',
        r'\botto\b',r'\byamazaki\b',r'\bomiki\b',
        r'\bvé\s*(ăn|buffet)\b',
    ]),
]

HIGH_RISK = [
    (r'\byogurt\b|\bsữa\s*chua\b',          'Yogurt ăn hay uống?'),
    (r'\bkem\b(?!\s*(que|hộp|lạnh|cone))',   'Kem: ice cream hay topping?'),
    (r'\bnước\s*yến\b|\byến\s*sào\b',        'Yến: thực phẩm hay đồ uống?'),
    (r'\bcombo\b|\bset\b',                    'Combo: có thể gồm cả đồ uống'),
    (r'\bbuffet\b',                           'Buffet: đồ ăn hay dịch vụ?'),
    (r'\bphô\s*mai\b|\bcheese\b',             'Phô mai: ăn hay nguyên liệu?'),
    (r'\bsốt\b|\bsauce\b',                    'Sốt: gia vị hay đồ ăn kèm?'),
    (r'\btopping\b|\bextra\b',               'Nguyên liệu topping, không phải món chính'),
    (r'\btrứng\s*gà\s*non\b',                'Trứng gà non: snack hay thực phẩm?'),
]

NOISE_PATTERNS = [
    (r'\bvé\b(?!\s*(ăn|buffet))',             'Vé - có thể là dịch vụ'),
    (r'^\s*\d+[kK]\b',                        'Giá tiền - không phải sản phẩm'),
    (r'\btopping\b|\bextra\b|\bupsize\b',     'Nguyên liệu/Topping'),
    (r'\bset\s*rau\s*ăn\s*lẩu\b',            'Combo rau - gần với Rau củ'),
]

def rule_classify(text):
    if not text.strip(): return 'Khác', 0.0, ''
    tl = text.lower()
    for label, priority, patterns in sorted(RULES, key=lambda x: x[1]):
        for pat in patterns:
            try:
                if re.search(pat, tl, re.IGNORECASE):
                    conf = 1.0 if priority == 1 else (0.85 if priority == 2 else (0.7 if priority == 3 else 0.6))
                    return label, conf, pat
            except re.error:
                continue
    return 'Khác', 0.0, ''

def flag_risk(text):
    tl = (text or '').lower()
    for pat, reason in HIGH_RISK:
        try:
            if re.search(pat, tl, re.IGNORECASE): return reason
        except: pass
    return None

def detect_noise(text):
    tl = (text or '').lower()
    for pat, reason in NOISE_PATTERNS:
        try:
            if re.search(pat, tl, re.IGNORECASE): return reason
        except: pass
    return None

print("⏳ Gán nhãn rule-based...")
res = df['clean'].apply(rule_classify)
df['rule_label'] = [r[0] for r in res]
df['rule_conf']  = [r[1] for r in res]
df['rule_match'] = [r[2] for r in res]
df['flag_reason'] = df['item_name'].apply(flag_risk)
df['noise_flag']  = df['item_name'].apply(detect_noise)
df['flag_review'] = (df['flag_reason'].notna() | (df['rule_conf'] < 0.5)).astype(int)

print("\nPhân phối rule_label:")
print(df['rule_label'].value_counts().to_string())
n_flag = df['flag_review'].sum()
print(f"\n🚩 Cần rà soát: {n_flag:,} ({n_flag/len(df)*100:.1f}%)")

df_lbl = df[df['label_ground'] != 'Khác']
match = (df_lbl['rule_label'] == df_lbl['label_ground']).mean()
print(f"✅ Khớp nhãn gốc: {match*100:.1f}% trên {len(df_lbl):,} mẫu")

print("\nTop 20 sai nhãn:")
mis = df_lbl[df_lbl['rule_label'] != df_lbl['label_ground']][
    ['item_name','label_ground','rule_label','rule_conf']].head(20)
print(mis.to_string(index=False))

# ─────────────────────────────────────────────────────────────
# BƯỚC 2: CHUẨN HÓA + FINAL LABEL
# ─────────────────────────────────────────────────────────────
print("\n"+"="*60); print("🔍 BƯỚC 2: CHUẨN HÓA & FINAL LABEL"); print("="*60)

def get_final_label(row):
    if row['label_ground'] not in ['Khác', '']:
        return row['label_ground'], 'ground-truth'
    if row['noise_flag'] is not None:
        return 'Loại bỏ', 'noise'
    if row['rule_conf'] >= 0.6:
        return row['rule_label'], 'rule-based'
    return 'Cần rà soát', 'uncertain'

fl = df.apply(get_final_label, axis=1)
df['final_label']  = [r[0] for r in fl]
df['label_source'] = [r[1] for r in fl]

print("Final label:")
print(df['final_label'].value_counts().to_string())
print("\nlabel_source:")
print(df['label_source'].value_counts().to_string())

# ─────────────────────────────────────────────────────────────
# BƯỚC 3: TRAIN MODEL
# ─────────────────────────────────────────────────────────────
print("\n"+"="*60); print("🤖 BƯỚC 3: TRAIN MODEL"); print("="*60)

df_tr = df[df['label_source'] == 'ground-truth'].copy()
valid = df_tr['final_label'].value_counts()
valid = valid[valid >= 5].index
df_tr = df_tr[df_tr['final_label'].isin(valid)]
print(f"Train: {len(df_tr):,} rows | {df_tr['final_label'].nunique()} classes")
print(df_tr['final_label'].value_counts().to_string())

le = LabelEncoder()
y  = le.fit_transform(df_tr['final_label'])
X  = df_tr['clean'].fillna('')
X_tr, X_te, y_tr, y_te = train_test_split(X, y, test_size=0.2, random_state=42, stratify=y)
print(f"\nTrain: {len(X_tr):,} | Test: {len(X_te):,}")

# Model A: char TF-IDF
print("\n⏳ Model A — char TF-IDF + LR...")
pipe_a = Pipeline([
    ('tfidf', TfidfVectorizer(ngram_range=(2,4), min_df=1,
                               max_features=20000, sublinear_tf=True, analyzer='char_wb')),
    ('clf', LogisticRegression(max_iter=1000, C=5.0, class_weight='balanced'))
])
pipe_a.fit(X_tr, y_tr)
f1_a = f1_score(y_te, pipe_a.predict(X_te), average='macro')
print(f"  F1-macro: {f1_a*100:.2f}%")

# Model B: word TF-IDF
print("⏳ Model B — word TF-IDF + LR...")
pipe_b = Pipeline([
    ('tfidf', TfidfVectorizer(ngram_range=(1,2), min_df=2,
                               max_features=20000, sublinear_tf=True)),
    ('clf', LogisticRegression(max_iter=1000, C=5.0, class_weight='balanced'))
])
pipe_b.fit(X_tr, y_tr)
f1_b = f1_score(y_te, pipe_b.predict(X_te), average='macro')
print(f"  F1-macro: {f1_b*100:.2f}%")

# Model C: XGBoost word
print("⏳ Model C — XGBoost...")
from sklearn.feature_extraction.text import TfidfVectorizer as TV
tv = TV(ngram_range=(1,2), min_df=2, max_features=20000, sublinear_tf=True)
Xtr_t = tv.fit_transform(X_tr); Xte_t = tv.transform(X_te)
xgb_clf = xgb.XGBClassifier(n_estimators=300, max_depth=6, learning_rate=0.1,
                              eval_metric='mlogloss', random_state=42, n_jobs=-1)
xgb_clf.fit(Xtr_t, y_tr)
f1_c = f1_score(y_te, xgb_clf.predict(Xte_t), average='macro')
print(f"  F1-macro: {f1_c*100:.2f}%")

f1s = [f1_a, f1_b, f1_c]; names = ['char TF-IDF+LR','word TF-IDF+LR','XGBoost']
best_idx = int(np.argmax(f1s)); best_f1 = f1s[best_idx]; best_name = names[best_idx]
print(f"\n{'='*50}")
for n, f in zip(names, f1s):
    mark = ' ← BEST' if n==best_name else ''
    print(f"  {n:<25}: {f*100:.2f}%{mark}")
print(f"{'='*50}")

best_pred = [pipe_a.predict(X_te), pipe_b.predict(X_te), xgb_clf.predict(Xte_t)][best_idx]
print(f"\n📋 Classification Report ({best_name}):")
print(classification_report(y_te, best_pred, target_names=le.classes_, zero_division=0))

f1_per = f1_score(y_te, best_pred, average=None)
print("Per-class F1:")
for cls, sc in sorted(zip(le.classes_, f1_per), key=lambda x: -x[1]):
    bar = '█'*int(sc*20)+'░'*(20-int(sc*20))
    print(f"  {cls:<35} {bar} {sc*100:5.1f}%")

# ─────────────────────────────────────────────────────────────
# PREDICT ALL + EXPORT
# ─────────────────────────────────────────────────────────────
print("\n"+"="*60); print("🔮 PREDICT ALL & EXPORT"); print("="*60)

X_all = df['clean'].fillna('')
if best_idx == 0:   prob = pipe_a.predict_proba(X_all)
elif best_idx == 1: prob = pipe_b.predict_proba(X_all)
else:               prob = xgb_clf.predict_proba(tv.transform(X_all))
pred = np.argmax(prob, axis=1); conf = np.max(prob, axis=1)
df['ml_label'] = le.inverse_transform(pred)
df['ml_conf']  = np.round(conf, 3)

def decide(row):
    if row['label_source'] == 'ground-truth': return row['final_label'], 'ground-truth'
    if row['noise_flag'] is not None:          return 'Loại bỏ', 'noise'
    if row['ml_conf'] >= 0.90: return row['ml_label'], 'ml-high'
    if row['ml_conf'] >= 0.60: return row['ml_label'], 'ml-medium'
    return 'Cần rà soát', 'uncertain'

dec = df.apply(decide, axis=1)
df['pred_label']  = [r[0] for r in dec]
df['pred_source'] = [r[1] for r in dec]

print("✅ Pred label:"); print(df['pred_label'].value_counts().to_string())
print("\n✅ Pred source:"); print(df['pred_source'].value_counts().to_string())

cols = ['category','user_id','product_group','item_name','clean',
        'label_ground','rule_label','rule_conf','rule_match',
        'ml_label','ml_conf','pred_label','pred_source',
        'flag_review','flag_reason','noise_flag']

df[cols].to_csv(OUT+'food_full_labeled.csv', index=False, encoding='utf-8-sig')
print(f"\n✅ food_full_labeled.csv ({len(df):,} rows)")

df_rft = df[df['pred_source'].isin(['ground-truth','ml-high']) & (df['pred_label'] != 'Loại bỏ')]
df_rft[['item_name','clean','pred_label']].to_csv(OUT+'food_ready_for_train.csv', index=False, encoding='utf-8-sig')
print(f"✅ food_ready_for_train.csv ({len(df_rft):,} rows)")

df_rev = df[df['pred_source'].isin(['uncertain','ml-medium'])][cols]
df_rev.sort_values(['flag_reason','ml_conf'], na_position='last').to_csv(OUT+'food_review_queue.csv', index=False, encoding='utf-8-sig')
print(f"✅ food_review_queue.csv ({len(df_rev):,} rows)")

rules_rows = []
for label, prio, patterns in RULES:
    for p in patterns:
        rules_rows.append({'label':label,'priority':prio,'pattern':p})
pd.DataFrame(rules_rows).to_csv(OUT+'food_keyword_rules_v1.csv', index=False, encoding='utf-8-sig')
print(f"✅ food_keyword_rules_v1.csv ({len(rules_rows)} rules)")

summary = {
    'total_items': int(len(df)),
    'ground_truth_labels': int((df['label_source']=='ground-truth').sum()),
    'rule_match_rate': f"{match*100:.1f}%",
    'best_model': best_name,
    'f1_char_lr': f"{f1_a*100:.2f}%",
    'f1_word_lr': f"{f1_b*100:.2f}%",
    'f1_xgb':     f"{f1_c*100:.2f}%",
    'best_f1_macro': f"{best_f1*100:.2f}%",
    'ready_for_train': int(len(df_rft)),
    'review_queue': int(len(df_rev)),
    'classes': list(le.classes_),
}
with open(OUT+'food_summary.json','w',encoding='utf-8') as f:
    json.dump(summary, f, ensure_ascii=False, indent=2)

print(f"""
╔══════════════════════════════════════════════════════════╗
║           FOOD PIPELINE — SUMMARY                       ║
╠══════════════════════════════════════════════════════════╣
║  Tổng items            : {len(df):>8,}                   ║
║  Ground-truth labels   : {(df['label_source']=='ground-truth').sum():>8,}                   ║
║  Rule match rate        : {match*100:>7.1f}%                   ║
║  char TF-IDF + LR      : {f1_a*100:>7.2f}%                   ║
║  word TF-IDF + LR      : {f1_b*100:>7.2f}%                   ║
║  XGBoost               : {f1_c*100:>7.2f}%                   ║
║  Best model            : {best_name:<28} ║
║  Best F1-macro         : {best_f1*100:>7.2f}%                   ║
║  Ready-for-train       : {len(df_rft):>8,}                   ║
║  Review queue          : {len(df_rev):>8,}                   ║
╚══════════════════════════════════════════════════════════╝
""")